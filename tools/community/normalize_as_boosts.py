"""Build src/data/attackSpeedBoosts.json from the community Attack Speed
Calculator (docs/Attack Speed Calculator.xlsx).

Two outputs:
  - globalItems: the item/emblem/ally attack-speed boosts (rows 1-10 of the
    'Atk Speed Boosts' sheet), each tagged active/passive/ally. Passive ones
    (Muscle Band, Choice Scarf, Red emblems) are already folded into the bundle
    via held-item stats + the red set bonus, so the UI must NOT toggle them
    (avoids double counting). Active ones (X-Attack, Rapid-Fire Scarf proc) and
    ally buffs are the real toggles.
  - moves: per-Pokémon move/ability AS boosts, parsed from the calculator's
    "Additional Attack Speed" formula, including level availability (minLevel/
    maxLevel) and per-level scaling where present.

Usage:  python3 normalize_as_boosts.py
"""

from __future__ import annotations

import json
import re
from pathlib import Path

import openpyxl
from openpyxl.worksheet.formula import ArrayFormula

HERE = Path(__file__).resolve().parent
XLSX = HERE.parent.parent / "docs" / "Attack Speed Calculator.xlsx"
OUT = HERE.parent.parent / "src" / "data" / "attackSpeedBoosts.json"


def slug(s: str) -> str:
    return "".join(c if c.isalnum() else "-" for c in s.lower()).strip("-")


# Semantic classification of the 10 global boosts (sheet rows 1-10). "passive"
# values already live in the bundle (item stats / red set bonus) -> not toggles.
GLOBAL_KIND = {
    "X-Atk": ("active", "x-attack"),
    "Muscle Band": ("passive", "muscle-band"),
    "Rapid Fire Scarf": ("passive", "rapid-fire-scarf"),
    "RFS Proc": ("active", "rapid-fire-scarf-proc"),
    "Choice Scarf": ("passive", "choice-scarf"),
    "Red 3": ("passive", "red-3"),
    "Red 5": ("passive", "red-5"),
    "Red 7": ("passive", "red-7"),
    "Blissey Helping Hand": ("ally", "blissey-helping-hand"),
    "Mew Coaching": ("ally", "mew-coaching"),
}


def parse_level_cond(cond: str | None):
    """Map an Excel level condition (on cell B4 = level) to min/max level."""
    out: dict[str, int] = {}
    if not cond:
        return out
    for m in re.finditer(r"B4\s*(<|>)\s*(\d+)", cond):
        op, n = m.group(1), int(m.group(2))
        if op == ">":
            out["minLevel"] = n + 1
        else:
            out["maxLevel"] = n - 1
    return out


def main() -> None:
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    wbf = openpyxl.load_workbook(XLSX, data_only=False)
    boosts = wb["Atk Speed Boosts"]

    # row -> (value, perLevel) from the boosts sheet
    row_value = {r: boosts.cell(r, 3).value for r in range(1, boosts.max_row + 1)}
    row_per = {r: boosts.cell(r, 4).value for r in range(1, boosts.max_row + 1)}

    # --- global items (rows 1-10) ---
    global_items = []
    for r in range(1, 11):
        name = boosts.cell(r, 2).value
        if name not in GLOBAL_KIND:
            continue
        kind, ident = GLOBAL_KIND[name]
        global_items.append({
            "id": ident,
            "label": name,
            "asPoints": row_value[r],
            "kind": kind,  # active | passive | ally
        })

    # --- per-Pokémon move boosts: parse the "Additional Attack Speed" formula ---
    calc = wbf["Atk Speed Calc"]
    cell = calc.cell(6, 2).value  # B6
    formula = cell.text if isinstance(cell, ArrayFormula) else str(cell)

    # Most clauses end `),'Atk Speed Boosts'!$C$N`. A few wrap the value in a
    # leading paren (e.g. Tsareena: `),('...$C$123+...$D$123*(B4-1))*$L$9`), so
    # allow an optional `(` before the row ref — perLevel still comes from col D.
    clause = re.compile(
        r'\$A\$4="([^"]+)",\$I\$\d+="([^"]+)",\$K\$\d+=TRUE'
        r"(?:,((?:OR\([^)]*\)|[^)])*?))?\),\(?'Atk Speed Boosts'!\$C\$(\d+)"
    )
    moves: dict[str, list] = {}
    for pokemon, source, cond, row in clause.findall(formula):
        row = int(row)
        entry = {"source": source, "asPoints": row_value.get(row)}
        per = row_per.get(row)
        if isinstance(per, (int, float)):
            entry["perLevel"] = per
        entry.update(parse_level_cond(cond))
        moves.setdefault(pokemon, []).append(entry)

    out = {
        "_source": "docs/Attack Speed Calculator.xlsx",
        "_note": "Attack-speed boosts in percentage points. 'passive' globals are "
                 "already in the bundle (item stats / red set bonus) — do not toggle them.",
        "globalItems": global_items,
        "moves": moves,
    }
    OUT.write_text(json.dumps(out, indent=2, ensure_ascii=False))
    print(f"Wrote {OUT}")
    print(f"  globalItems={len(global_items)} | pokemon with move boosts={len(moves)} | "
          f"total move boosts={sum(len(v) for v in moves.values())}")


if __name__ == "__main__":
    main()
